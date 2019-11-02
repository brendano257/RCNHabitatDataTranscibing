"""
Proof of concept that reads and outputs to a new file/format.

Data are taken along a 50-meter transect at height ranges.
Collection sheets have up three species listed per height range, with one possible overflow value.
Each sheet has a date, transect name, and hard-coded site 'common' name that applies to all points.
"""
from pathlib import Path
from openpyxl import load_workbook, Workbook
from collections import defaultdict

CORE_DIR = Path('/home/brendan/PycharmProjects/RCNHabitatDataTranscibing')
DATA_DIR = CORE_DIR / 'data'
CREATED_DIR = CORE_DIR / 'created'

INPUT_FILE = DATA_DIR / 'Line-point Intercept Data RNC 2019.xlsx'
OUTPUT_FILE = CREATED_DIR / 'new_output_file.xlsx'

START_ROW = 4

HEADER_LINE = ['transect_ID',	'sample_date',	'sample_site',	'sample_point',	'stratum',	'species']

book = load_workbook(INPUT_FILE)
all_new_lines = [HEADER_LINE]

for sheetname in book.sheetnames:
    worksheet = book[sheetname]
    date = worksheet.cell(3, 4).value
    transect = worksheet.cell(2, 4).value

    for index, row in enumerate(worksheet.rows):
        strata = defaultdict(list)

        point = row[0].value

        if index > START_ROW:
            strata['0-1'].extend([cell.value for cell in row[2:5]])
            strata['1-2'].extend([cell.value for cell in row[5:8]])
            strata['2-5'].extend([cell.value for cell in row[8:11]])
            strata['5 +'].extend([cell.value for cell in row[11:14]])

            # index the overflow as {value: species}, eg {'0-1': DANSPI}
            overflow = {row[14].value: row[15].value}

            for stratum_name, stratum in strata.items():
                for species in stratum:
                    if species:
                        row_output = [transect, date.strftime('%Y-%m-%d'),
                                      "Concord Pine Barrens", point, stratum_name, species]
                        all_new_lines.append(row_output)

                overflow_value = overflow.get(stratum_name)

                if overflow_value:
                    overflow_output = [transect, date.strftime('%Y-%m-%d'),
                                  "Concord Pine Barrens", point, stratum_name, overflow_value]

                    all_new_lines.append(overflow_output)

workbook = Workbook()
worksheet = workbook.active

from openpyxl.styles import Alignment, Font
al = Alignment(horizontal='centerContinuous')

for row in all_new_lines:
    worksheet.append(row)

for col in worksheet.columns:
    for cell in col:
        cell.alignment = al

for cell in next(worksheet.rows):
    cell.font = Font(name='Calibri', size=12, bold=True)

workbook.save(OUTPUT_FILE)
